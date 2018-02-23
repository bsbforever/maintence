#coding=utf8


from ping3 import ping, verbose_ping
import socket
from openpyxl import Workbook
from openpyxl import load_workbook


def ifping(ipaddress): #检查IP地址是否连通
    result=ping( ipaddress)
    return result

def  checkport(ipaddress,port): #检查端口是否开放
    sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    result = sock.connect_ex((ipaddress, port))
    return result




if __name__=="__main__":
    #ipaddresses=['10.65.206.23','10.65.1.168','10.65.202.84']
    dbs=['1521:oracle','1522:oracle','1523:oracle','1525:oracle','1527:oracle','1529:oracle','1433:sqlserver','3306:mysql'] # 将需要检查的数据库端口放入列表
    networks=['10.65.202'] #将需要检查的网段放入列表
    path=r'x:\dbscan.xlsx'
    wb = load_workbook(path) #打开excel文件，需要事先存在

    # 初始化第一个sheet
    #work_sheet = wb.active
    #work_sheet.title = "range names"
    for network in networks:
        ## 新建sheet用于存放各个网段并初始化
        if  network in wb.sheetnames: #wb.sheetnames为excel的所有sheet名称
            work_sheet = wb.get_sheet_by_name(network) #指定数据库信息写入的sheet名称
        else:
            wb.create_sheet(title=network) #新建sheet，title指定sheet名称
            work_sheet = wb.get_sheet_by_name(network)
            work_sheet.append(['IP地址', '数据库']) #在新建sheet第一行设置标题
            wb.save(path) #保存文件

        for i in range(1,255):
            ipaddress=network+'.'+str(i)
            print ('now checking '+ipaddress)
            result = ifping(ipaddress)
            if result is  None:
                #print(ipaddress + ' is  not reachable')
                work_sheet.append([ipaddress, 'not reachable']) #如不通则写入相关信息至excel
                wb.save(path)

            else:
                #print(ipaddress + ' is  reachable')
                opened = 0  # 用于判断是否无DB端口开放
                for db in dbs:
                    port=int(db.split(':')[0])
                    name=db.split(':')[1]
                    ifopen=checkport(ipaddress, port)
                    #print (result)
                    if ifopen==0:
                        #print (ipaddress +' have '+name +' installed')
                        opened=opened+1
                        work_sheet.append([ipaddress,name]) #写入开放的数据库端口
                        wb.save(path)
                if opened==0:
                    work_sheet.append([ipaddress, 'No DBs Found']) #如没有端口开放则写入未找到数据库
                    wb.save(path)
    #wb.save(path)
    wb.close() #关闭文件
